# import load_workbook
import json
from email_validator import validate_email, EmailNotValidError
from openpyxl import load_workbook
from datetime import datetime


#filepath=C:\Users\Jeffrey PC\Downloads\Untitled spreadsheet (3).xlsx
flaggedRows=[]
courseCategories=[]

def main():
    path=input("Spreadsheet Directory: ")
    dataType=input("Input data type")
    checkSheet(path,dataType)

def checkSheet(path,dataType):
    sheet = load_workbook(path).active
    for currentRow in range(2, sheet.max_row + 1):
        for currentColumn in range(1, sheet.max_column + 1):
            currentValue = sheet.cell(row=currentRow, column=currentColumn).value
            if(dataType=="Student"):
                print(sheet.cell(row=currentRow, column=currentColumn).value)
                if(currentColumn in [1,2,3,5,6]):
                    if(nameChecker(currentValue)==False):
                        if(currentColumn<=2):
                            flaggedRows.append([currentRow, "student name error"])
                        if(currentColumn==3):
                            flaggedRows.append([currentRow, "school name error"])
                        if(currentColumn>=5):
                            flaggedRows.append([currentRow, "parent name error"])
                if(currentColumn == 4):
                    print(currentValue)
                    print(currentValue.strftime('%m/%d/%Y'))
                    try:
                        (datetime.strptime(currentValue.strftime('%m/%d/%Y'), '%m/%d/%Y'))
                        print(currentValue)
                    except ValueError:
                        flaggedRows.append([currentRow, "date format error"])
                if(currentColumn == 7):
                    if(currentColumn!=emailValidator(currentValue)):
                        flaggedRows.append([currentRow, emailValidator(currentValue)])
                if(currentColumn == 8):
                    if(currentValue.is_integer()==False):
                        flaggedRows.append([currentRow,"Grade error"])
            if(dataType=="Instructor"):
                if(currentColumn in [1,2]):
                    if(nameChecker(currentValue)==False):
                        if(currentColumn<=2):
                            flaggedRows.append([currentRow, "Instructor name error"])
                if(currentColumn == 3):
                    if (currentColumn != emailValidator(currentValue)):
                        flaggedRows.append([currentRow, emailValidator(currentValue)])
                if (currentColumn == 8):
                    if (currentValue.is_integer() == False):
                        flaggedRows.append([currentRow, "Years of Experience Error"])
            if(dataType=="Course"):
                if(currentColumn in [1]):
                    #check if course exists
                    if(nameChecker(currentValue)==False):
                        flaggedRows.append([currentRow, "course name error"])
                if(currentColumn in [2,3]):
                    dateValidator()





    print(flaggedRows)

def dateTimeValidator(date, format):
    try:
        datetime.strptime(date, format)
    except ValueError:
        print("sdf")

def nameChecker(name):
    if name.replace(" ","").isalpha():
        return True
    else:
        return False

def emailValidator(email):
    try:
        validate_email(email).email
    except EmailNotValidError as e:
        # email is not valid, exception message is human-readable
        return str(e)

def dateValidator():
    print("date validator")

if __name__ == '__main__':
    main()